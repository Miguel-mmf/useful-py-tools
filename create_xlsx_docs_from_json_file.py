import json
from collections import defaultdict

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from tqdm import tqdm

FILE_PATH: str = 'data/input_model.json'


def explore_json(data: dict, parent_key: str = '') -> list:
    """Explores a JSON structure and extracts key-value pairs.

    This function recursively navigates a dictionary representing a JSON
    structure, flattening it into a list of tuples. Each tuple contains
    the full key path, the type of the value, and the value itself.
    It handles nested dictionaries and lists, extracting information
    from each level.

    Args:
        data (dict): The JSON data as a dictionary.
        parent_key (str, optional): The key path of the parent object. Defaults to "".

    Returns:
        list: A list of tuples, where each tuple contains the key path,
            value type, and example value.
    """
    items = []
    for key, value in data.items():
        new_key = f'{parent_key}.{key}' if parent_key else key
        if isinstance(value, dict):
            items.extend(explore_json(value, new_key))
        elif isinstance(value, list):
            if value:
                first_value = value[0]
                if isinstance(first_value, dict):
                    items.extend(explore_json(first_value, f'{new_key}'))
                elif isinstance(first_value, str):
                    items.append((new_key, 'list', first_value))
        else:
            items.append((new_key, type(value).__name__, value))
    return items


def read_json_file(file_path: str) -> dict:
    """Reads a JSON file and returns its content as a dictionary.

    This function opens the specified JSON file, reads its contents,
    and parses it into a Python dictionary using the `json` module.
    It handles UTF-8 encoding to support a wide range of characters.

    Args:
        file_path (str): The path to the JSON file.

    Returns:
        dict: The JSON data as a Python dictionary.
    """
    with open(file_path, 'r', encoding='utf-8') as file:
        data = json.load(file)
    return data


def create_xlsx_docs(file_path: str = FILE_PATH) -> None:
    """Creates XLSX documentation from a JSON file.

    This function reads a JSON file, extracts its structure,
    and generates an XLSX file containing documentation tables.
    Each table represents a main key in the JSON data, with columns
    for key paths, value types, examples, and additional information.

    Args:
        file_path (str, optional): The path to the JSON file.
            Defaults to FILE_PATH.

    Returns:
        None
    """
    data = read_json_file(file_path)

    if not data:
        raise ValueError('O arquivo JSON está vazio!')

    # if 'output' in file_path.split('/')[-1]:
    #     content = data.get('result', {})
    # elif 'input' in file_path.split('/')[-1]:
    #     content = data.get('content', {})
    # else:
    #     raise ValueError(
    #         "O arquivo JSON não possui a chave 'content' ou 'result'!"
    #     )

    tables = defaultdict(list)
    main_keys_with_types = set()
    for main_key, sub_dict in tqdm(data.items()):
        tqdm.write(f'Processando {main_key}...', end='\r')
        main_keys_with_types.add((main_key, type(sub_dict).__name__))
        if isinstance(sub_dict, dict):
            records = explore_json(sub_dict)
            for record in records:
                keys_path, value_type, example_value = record
                keys_split = keys_path.split('.')
                depth = len(keys_split)
                row = {
                    f'Chave nível {i + 1}': keys_split[i] for i in range(depth)
                }
                row['Tipo'] = value_type
                row['Exemplo'] = example_value
                tables[main_key].append(row)
        elif isinstance(sub_dict, list) and sub_dict:
            first_value = sub_dict[0]
            records = explore_json(first_value)
            for record in records:
                keys_path, value_type, example_value = record
                keys_split = keys_path.split('.')
                depth = len(keys_split)
                row = {
                    f'Chave nível {i + 1}': keys_split[i] for i in range(depth)
                }
                row['Tipo'] = value_type
                row['Exemplo'] = example_value
                tables[main_key].append(row)

        tqdm.write(f'Processando {main_key}... OK')

    output_filename = file_path.replace('.json', '.xlsx')
    with pd.ExcelWriter(output_filename) as writer:
        
        df = pd.DataFrame()
        df['Chave primária'] = [main_key[0] for main_key in main_keys_with_types]
        df['Tipo'] = [main_key[1] for main_key in main_keys_with_types]
        df['Significado'] = None
        df['Observações'] = None
        df.fillna('---', inplace=True)
        df.to_excel(writer, sheet_name='Chaves Principais', index=False)
        
        for sheet_name, data in tables.items():
            df = pd.DataFrame(data)
            df['Significado'] = None
            df['Unidade'] = None
            df['Obrigatório'] = 'SIM'
            df['Limite Mínimo'] = None
            df['Limite Máximo'] = None
            df['Observações'] = None

            # Reorder columns
            columns = [
                col for col in df.columns.tolist() if col.startswith('Chave')
            ]
            columns = sorted(columns, key=lambda x: x.split()[-1])
            columns.extend([
                'Exemplo',
                'Tipo',
                'Unidade',
                'Significado',
                'Obrigatório',
                'Observações',
                'Limite Mínimo',
                'Limite Máximo',
            ])
            df = df[columns]

            # rename columns
            df.rename(
                columns={
                    'Chave nível 1': 'Chave primária',
                    'Chave nível 2': 'Chave secundária',
                    'Chave nível 3': 'Chave terciária',
                    'Chave nível 4': 'Chave quaternária',
                    'Chave nível 5': 'Chave quinária',
                    'Chave nível 6': 'Chave senária',
                },
                inplace=True,
                errors='ignore',
            )

            # fill NaN values
            df.fillna('---', inplace=True)

            df = (
                df.groupby([
                    col
                    for col in df.columns[:3].tolist()
                    if col.startswith('Chave')
                ])
                .agg({
                    'Exemplo': 'first',
                    'Tipo': 'first',
                    'Unidade': 'first',
                    'Significado': 'first',
                    'Obrigatório': 'first',
                    'Observações': 'first',
                    'Limite Mínimo': 'first',
                    'Limite Máximo': 'first',
                })
                .reset_index()
            )
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    config_xlsx(output_filename)
    print('Arquivo Excel gerado com sucesso!')


def merge_equal_cells(sheet, col: str) -> None:
    """Merges cells with equal values in a given column of an Excel sheet.

    This function iterates through the specified column of the provided
    `openpyxl` worksheet and merges contiguous cells that have the same
    value. It starts from the first row and continues until the last row,
    merging cells vertically whenever a sequence of identical values is found.

    Args:
        sheet: The `openpyxl` worksheet object.
        col: The column letter (e.g., 'A', 'B', 'C') to process.

    Returns:
        None
    """
    max_row = sheet.max_row
    start_row = 1
    while start_row <= max_row:
        cell_value = sheet[f'{col}{start_row}'].value
        end_row = start_row
        while (
            end_row <= max_row and sheet[f'{col}{end_row}'].value == cell_value
        ):
            end_row += 1
        if end_row - start_row > 1:
            sheet.merge_cells(f'{col}{start_row}:{col}{end_row - 1}')
        start_row = end_row


def config_xlsx(xlsx_path: str) -> None:
    """Configures and styles an XLSX file.

    This function takes the path to an XLSX file, loads it using `openpyxl`,
    and applies various formatting styles. It sets fonts, fills, borders,
    merges cells, and configures auto-filters to enhance the readability
    and presentation of the data in the Excel sheets.

    Args:
        xlsx_path (str): The path to the XLSX file.

    Returns:
        None
    """
    wb = load_workbook(xlsx_path)

    # Definir as cores de preenchimento
    fill_yes = PatternFill(
        start_color='00FF00', end_color='00FF00', fill_type='solid'
    )  # Verde para "Sim"
    fill_no = PatternFill(
        start_color='FF0000', end_color='FF0000', fill_type='solid'
    )  # Vermelho para "Não"
    mapping_color_per_type = {
        'int': 'FFFF00',  # Amarelo
        'float': 'FFA500',  # Laranja
        'str': 'ADD8E6',  # Azul Claro
        'list': '90EE90',  # Verde Claro
        'dict': 'D3D3D3',  # Cinza Claro
        'bool': 'FFC0CB',  # Rosa Claro
    }
    mapping_color_per_type = {
        k: PatternFill(start_color=v, end_color=v, fill_type='solid')
        for k, v in mapping_color_per_type.items()
    }
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = Font(bold=True)

        for row in ws.iter_rows(
            min_row=2, max_row=ws.max_row, min_col=1, max_col=1
        ):
            for cell in row:
                if cell.value.lower() == 'sim':
                    cell.fill = fill_yes
                elif cell.value.lower() == 'não':
                    cell.fill

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
                cell.alignment = Alignment(
                    horizontal='center', vertical='center'
                )
                cell.border = thin_border
                if (
                    col[0].value == 'Tipo'
                    and cell.value in mapping_color_per_type
                ):
                    cell.fill = mapping_color_per_type.get(cell.value)

            adjusted_width = max_length + 8
            ws.column_dimensions[column].width = adjusted_width

            if col[0].value.startswith('Chave'):
                merge_equal_cells(ws, column)

            if col[0].value == 'Obrigatório':
                for cell in col:
                    if cell.value.lower() == 'sim':
                        cell.fill = fill_yes
                    elif cell.value.lower() == 'não':
                        cell.fill = fill_no

        ws.auto_filter.ref = ws.dimensions

    wb.save(xlsx_path)


if __name__ == '__main__':
    create_xlsx_docs()
